"""JobPilot Excel exporter — 4-sheet professional output."""

import io
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Styles ─────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="0A1628", end_color="0A1628", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
ROW_EVEN = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
ROW_ODD = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="DEE2E6"),
    right=Side(style="thin", color="DEE2E6"),
    top=Side(style="thin", color="DEE2E6"),
    bottom=Side(style="thin", color="DEE2E6"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")

STATUS_FILLS = {
    "✓ Open": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
    "⚠ Unverified": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
    "✗ Closed": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
}

SCORE_FILLS = {
    5: PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
    4: PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
    3: PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
    2: PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
    1: PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
    0: PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
}


def _write_header(ws, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = THIN_BORDER
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def _style_cell(ws, row, col, value, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = ROW_EVEN if row % 2 == 0 else ROW_ODD
    c.border = THIN_BORDER
    c.alignment = WRAP if wrap else Alignment(vertical="top")
    c.font = Font(size=10, name="Calibri")
    return c


def _auto_width(ws, min_w=12, max_w=60):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value or "")
            # Use first line only for width calc
            first_line = val.split("\n")[0] if "\n" in val else val
            max_len = max(max_len, len(first_line))
        ws.column_dimensions[col_letter].width = max(min_w, min(max_len + 3, max_w))


# ── Sheet 1: Jobs ──────────────────────────────────────────

def _build_jobs_sheet(wb, jobs):
    ws = wb.active
    ws.title = "Jobs"

    headers = ["#", "Company", "Title", "Location", "Remote", "Status",
               "Confidence", "Visa", "Degree", "CV Version", "Apply URL",
               "Job Board", "Posted", "Match Reason"]
    _write_header(ws, headers)

    for i, job in enumerate(jobs, 1):
        r = i + 1
        audit = job.get("audit", {})
        cl = job.get("cover_letter", {})

        _style_cell(ws, r, 1, i)
        _style_cell(ws, r, 2, job.get("company", ""))
        _style_cell(ws, r, 3, job.get("title", ""))
        _style_cell(ws, r, 4, job.get("location", ""))
        _style_cell(ws, r, 5, "Yes" if job.get("remote") else "No")

        # Status with conditional color
        status = audit.get("status", "⚠ Unverified")
        sc = _style_cell(ws, r, 6, status)
        sc.fill = STATUS_FILLS.get(status, ROW_ODD)
        sc.font = Font(size=10, bold=True, name="Calibri")

        _style_cell(ws, r, 7, audit.get("confidence", ""))
        _style_cell(ws, r, 8, audit.get("visa", ""))
        _style_cell(ws, r, 9, audit.get("degree_match", ""))
        _style_cell(ws, r, 10, job.get("recommended_cv", cl.get("recommended_cv", "")))

        url_cell = _style_cell(ws, r, 11, job.get("apply_url", ""))
        url_cell.font = Font(size=9, color="0563C1", underline="single", name="Calibri")

        _style_cell(ws, r, 12, job.get("job_board", ""))
        _style_cell(ws, r, 13, job.get("posting_date", audit.get("posting_age_days", "")))
        _style_cell(ws, r, 14, job.get("match_reason", audit.get("reason", "")), wrap=True)

        ws.row_dimensions[r].height = 40

    _auto_width(ws)


# ── Sheet 2: Cover Letters ─────────────────────────────────

def _build_cl_sheet(wb, jobs):
    ws = wb.create_sheet("Cover Letters")

    headers = ["#", "Company", "Title", "Cover Letter", "Words", "Score",
               "Gates Passed", "Needs Review", "Metrics Used"]
    _write_header(ws, headers)

    row = 2
    for i, job in enumerate(jobs, 1):
        cl = job.get("cover_letter", {})
        if not cl.get("text"):
            continue

        _style_cell(ws, row, 1, i)
        _style_cell(ws, row, 2, job.get("company", ""))
        _style_cell(ws, row, 3, job.get("title", ""))
        _style_cell(ws, row, 4, cl.get("text", ""), wrap=True)
        _style_cell(ws, row, 5, cl.get("word_count", 0))

        score = cl.get("score", 0)
        sc = _style_cell(ws, row, 6, f"{score}/5")
        sc.fill = SCORE_FILLS.get(score, ROW_ODD)
        sc.font = Font(size=10, bold=True, name="Calibri")

        gates = cl.get("gates", {})
        passed = [k for k, v in gates.items() if v]
        _style_cell(ws, row, 7, ", ".join(passed))
        _style_cell(ws, row, 8, "YES" if cl.get("needs_review") else "No")
        _style_cell(ws, row, 9, ", ".join(cl.get("metrics_used", [])))

        ws.row_dimensions[row].height = 200
        row += 1

    # Make CL column wider
    ws.column_dimensions["D"].width = 80
    _auto_width(ws, min_w=12, max_w=80)


# ── Sheet 3: Tracker ───────────────────────────────────────

def _build_tracker_sheet(wb, jobs):
    ws = wb.create_sheet("Tracker")

    headers = ["#", "Company", "Title", "Status", "Applied Date",
               "Response", "Interview Date", "Notes", "Deadline", "CV Used"]
    _write_header(ws, headers)

    for i, job in enumerate(jobs, 1):
        r = i + 1
        _style_cell(ws, r, 1, i)
        _style_cell(ws, r, 2, job.get("company", ""))
        _style_cell(ws, r, 3, job.get("title", ""))
        _style_cell(ws, r, 4, "Pending")
        _style_cell(ws, r, 5, "")  # Applied Date — user fills
        _style_cell(ws, r, 6, "")  # Response — user fills
        _style_cell(ws, r, 7, "")  # Interview Date
        _style_cell(ws, r, 8, "")  # Notes
        _style_cell(ws, r, 9, "")  # Deadline
        _style_cell(ws, r, 10, job.get("recommended_cv", ""))

        ws.row_dimensions[r].height = 25

    _auto_width(ws)


# ── Sheet 4: Audit Log ────────────────────────────────────

def _build_audit_sheet(wb, jobs):
    ws = wb.create_sheet("Audit Log")

    # Audit log header with explanation
    ws.merge_cells("A1:P1")
    intro = ws.cell(row=1, column=1,
        value="VERIFICATION AUDIT LOG — This sheet shows exactly how each job was verified. "
              "✓ Open = confirmed via ATS API or apply button found. "
              "⚠ Unverified = could not confirm (check manually). "
              "✗ Closed = definitive closed signal found.")
    intro.font = Font(size=9, italic=True, color="666666", name="Calibri")
    intro.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[1].height = 35

    headers = ["#", "Company", "Title", "Verification Method", "URL Checked",
               "Final URL", "HTTP", "Status", "Confidence", "Reasoning",
               "Posted (days ago)", "Ghost Risk", "Degree Match", "Visa Status",
               "Dropped?", "Drop Reason"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = THIN_BORDER
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"

    for i, job in enumerate(jobs, 1):
        r = i + 2  # Row 1 = intro, Row 2 = headers, data starts at 3
        a = job.get("audit", {})
        board = job.get("job_board", "direct")

        # Determine verification method label
        method = {
            "greenhouse": "Greenhouse API (definitive)",
            "lever": "Lever API (definitive)",
        }.get(board, f"HTML page analysis ({board})")

        _style_cell(ws, r, 1, i)
        _style_cell(ws, r, 2, job.get("company", ""))
        _style_cell(ws, r, 3, job.get("title", ""))
        _style_cell(ws, r, 4, method)
        _style_cell(ws, r, 5, a.get("url_checked", ""))

        final_url = a.get("final_url", "")
        _style_cell(ws, r, 6, final_url if final_url != a.get("url_checked") else "same")
        _style_cell(ws, r, 7, a.get("http_status", ""))

        status = a.get("status", "⚠ Unverified")
        sc = _style_cell(ws, r, 8, status)
        sc.fill = STATUS_FILLS.get(status, ROW_ODD)
        sc.font = Font(size=10, bold=True, name="Calibri")

        _style_cell(ws, r, 9, a.get("confidence", ""))
        _style_cell(ws, r, 10, a.get("reason", ""), wrap=True)
        _style_cell(ws, r, 11, a.get("posting_age_days", ""))

        ghost = a.get("ghost_risk", "unknown")
        gc = _style_cell(ws, r, 12, ghost)
        if ghost == "high":
            gc.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

        _style_cell(ws, r, 13, a.get("degree_match", ""))
        _style_cell(ws, r, 14, a.get("visa", ""))
        _style_cell(ws, r, 15, "YES" if a.get("drop") else "No")
        _style_cell(ws, r, 16, a.get("drop_reason", ""), wrap=True)

        ws.row_dimensions[r].height = 40

    _auto_width(ws)


# ── Main export function ──────────────────────────────────

def export_excel(jobs: list[dict], profile_name: str = "Candidate") -> bytes:
    """
    Export jobs + CLs + tracker + audit to a styled Excel file.
    Returns file bytes ready for HTTP response.
    """
    wb = openpyxl.Workbook()

    _build_jobs_sheet(wb, jobs)
    _build_cl_sheet(wb, jobs)
    _build_tracker_sheet(wb, jobs)
    _build_audit_sheet(wb, jobs)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
